#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
universidad_ranking_tracker.py

Actualiza un Excel de universidades con identificadores ROR y métricas académicas
obtenidas desde OpenAlex. Está pensado para usarse sobre el archivo:
    data/top_200_universidades_QS_2026_con_ARG_LATAM_recomendadas.xlsx

Qué hace:
- Lee una hoja del Excel con universidades o todas las hojas con --all-sheets.
- Detecta columnas de universidad, país y ranking QS.
- Busca el identificador ROR de cada universidad.
- Usa ese ROR para consultar la institución en OpenAlex.
- Agrega métricas: works_count, cited_by_count, h-index, i10, publicaciones/citas por año.
- Crea/actualiza las hojas:
    * Tracker_OpenAlex_ROR
    * Historial_metricas

Novedad v1.1:
- Agrega --all-sheets para procesar automáticamente todas las hojas que tengan columnas
  reconocibles de universidad. Salta hojas de notas/salida y elimina duplicados.

Importante:
- QS/THE/ARWU no tienen actualización realmente "en tiempo real" para puestos oficiales.
  Este script conserva el ranking QS del Excel y agrega métricas dinámicas de OpenAlex/ROR.
- OpenAlex requiere API key. Creala gratis y pasala con --openalex-api-key o variable OPENALEX_API_KEY.

Uso rápido:
    pip install -r requirements.txt
    set OPENALEX_API_KEY=TU_API_KEY   # Windows CMD
    # o: $env:OPENALEX_API_KEY="TU_API_KEY"  # PowerShell
    python universidad_ranking_tracker.py --input data/top_200_universidades_QS_2026_con_ARG_LATAM_recomendadas.xlsx --output universidades_tracker_actualizado.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import logging
import os
import re
import sys
import time
import unicodedata
import urllib.parse
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    import pycountry
except Exception:  # pragma: no cover
    pycountry = None


OPENALEX_BASE = "https://api.openalex.org"
ROR_BASE_V2 = "https://api.ror.org/v2/organizations"
ROR_BASE_V1 = "https://api.ror.org/organizations"

TRACKER_SHEET = "Tracker_OpenAlex_ROR"
HISTORY_SHEET = "Historial_metricas"
INSTITUTIONS_SHEET = "Instituciones"
RANKINGS_HISTORY_SHEET = "Rankings_Historicos"
OPENALEX_METRICS_SHEET = "Metricas_OpenAlex"
PENDING_MATCHES_SHEET = "Matches_Pendientes"
QUALITY_SHEET = "Control_Calidad"

SOURCE_SHEET_COL = "tracker_hoja_origen"
SOURCE_HEADER_ROW_COL = "tracker_fila_encabezado"
STD_NAME_COL = "tracker_universidad"
STD_COUNTRY_COL = "tracker_pais"
STD_RANK_COL = "tracker_ranking_base"
DEDUP_KEY_COL = "tracker_clave_dedupe"

RANKINGS_SHEET = "Rankings_Oficiales"
RANKINGS_SOURCES_SHEET = "Fuentes_Rankings"

# Campos útiles y livianos. Reducen tamaño de respuesta y consumo.
OPENALEX_SELECT = (
    "id,ror,display_name,country_code,type,homepage_url,works_count,"
    "cited_by_count,summary_stats,geo,ids,counts_by_year,works_api_url,updated_date"
)

COLUMN_ALIASES = {
    "name": [
        "universidad",
        "institution",
        "institution name",
        "university",
        "nombre universidad",
        "nombre",
    ],
    "country": [
        "pais/territorio qs",
        "pais",
        "country",
        "location",
        "pais/territorio",
    ],
    "rank": [
        "ranking qs 2026",
        "rank",
        "ranking",
        "qs rank",
        "rango qs 2026",
    ],
}

COUNTRY_NAME_FIXES = {
    "United States of America": "United States",
    "United Kingdom": "United Kingdom",
    "Hong Kong SAR": "Hong Kong",
    "Taiwan": "Taiwan",
    "Korea, South": "Korea, Republic of",
    "South Korea": "Korea, Republic of",
    "Russia": "Russian Federation",
    "Iran": "Iran, Islamic Republic of",
    "Viet Nam": "Vietnam",
    "Czech Republic": "Czechia",
    "Türkiye": "Turkey",
    "Mainland China": "China",
}


def now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


def setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.split())


def find_header_row(input_path: Path, sheet_name: str, max_rows: int = 25) -> int:
    preview = pd.read_excel(input_path, sheet_name=sheet_name, header=None, nrows=max_rows)
    aliases = {normalize_header(x) for values in COLUMN_ALIASES.values() for x in values}
    for idx, row in preview.iterrows():
        normalized_values = {normalize_header(v) for v in row.tolist() if str(v) != "nan"}
        if normalized_values & aliases:
            return int(idx) + 1  # Excel rows start at 1
    raise ValueError(
        f"No pude detectar la fila de encabezados en la hoja '{sheet_name}'. "
        "Usá --header-row, por ejemplo --header-row 4."
    )


def find_column(df: pd.DataFrame, kind: str, explicit: Optional[str] = None) -> str:
    if explicit:
        if explicit not in df.columns:
            raise ValueError(f"La columna indicada '{explicit}' no existe. Columnas: {list(df.columns)}")
        return explicit

    aliases = {normalize_header(a) for a in COLUMN_ALIASES[kind]}
    for col in df.columns:
        if normalize_header(col) in aliases:
            return col
    # Búsqueda parcial: "Universidad" dentro de "Nombre Universidad".
    for col in df.columns:
        col_norm = normalize_header(col)
        if any(alias in col_norm for alias in aliases):
            return col
    raise ValueError(f"No pude detectar la columna de tipo '{kind}'. Columnas: {list(df.columns)}")


def country_to_alpha2(country_name: Any) -> Optional[str]:
    if not country_name or str(country_name).lower() == "nan":
        return None
    raw = str(country_name).strip()
    fixed = COUNTRY_NAME_FIXES.get(raw, raw)

    if pycountry is None:
        return None
    try:
        return pycountry.countries.lookup(fixed).alpha_2
    except LookupError:
        # Algunos nombres del Excel pueden traer notas o territorios especiales.
        candidates = [c for c in pycountry.countries if fixed.lower() in c.name.lower()]
        if candidates:
            return candidates[0].alpha_2
    return None


def load_cache(cache_path: Path) -> Dict[str, Any]:
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            logging.warning("No pude leer la caché; se creará una nueva: %s", cache_path)
    return {"created_at": now_iso(), "items": {}}


def save_cache(cache_path: Path, cache: Dict[str, Any]) -> None:
    cache["updated_at"] = now_iso()
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def request_json(
    session: requests.Session,
    url: str,
    *,
    params: Optional[Dict[str, Any]] = None,
    headers: Optional[Dict[str, str]] = None,
    timeout: int = 40,
    retries: int = 3,
    sleep_base: float = 1.5,
) -> Optional[Dict[str, Any]]:
    for attempt in range(1, retries + 1):
        try:
            resp = session.get(url, params=params, headers=headers, timeout=timeout)
            if resp.status_code in {429, 500, 502, 503, 504} and attempt < retries:
                wait = sleep_base * attempt
                logging.warning("HTTP %s en %s. Reintento en %.1fs", resp.status_code, url, wait)
                time.sleep(wait)
                continue
            if resp.status_code == 404:
                return None
            resp.raise_for_status()
            return resp.json()
        except requests.RequestException as exc:
            if attempt >= retries:
                logging.warning("Fallo definitivo consultando %s: %s", url, exc)
                return None
            wait = sleep_base * attempt
            logging.warning("Fallo consultando %s: %s. Reintento en %.1fs", url, exc, wait)
            time.sleep(wait)
    return None


def extract_ror_name(org: Dict[str, Any]) -> Optional[str]:
    if org.get("name"):
        return org.get("name")
    names = org.get("names") or []
    if isinstance(names, list):
        # Schema 2.x: names = [{value, types:[ror_display/preferred/...]}]
        for preferred_type in ("ror_display", "preferred", "label"):
            for n in names:
                if isinstance(n, dict) and preferred_type in (n.get("types") or []):
                    return n.get("value")
        for n in names:
            if isinstance(n, dict) and n.get("value"):
                return n.get("value")
    return None


def extract_ror_country(org: Dict[str, Any]) -> Tuple[Optional[str], Optional[str]]:
    # Schema 1.x
    country = org.get("country")
    if isinstance(country, dict):
        return country.get("country_name"), country.get("country_code")

    # Schema 2.x
    locations = org.get("locations") or []
    if isinstance(locations, list) and locations:
        geonames = locations[0].get("geonames_details") or {}
        country_name = geonames.get("country_name") or geonames.get("country")
        country_code = geonames.get("country_code")
        return country_name, country_code
    return None, None


def extract_ror_lat_lon(org: Dict[str, Any]) -> Tuple[Optional[float], Optional[float]]:
    # Schema 1.x
    addresses = org.get("addresses") or []
    if addresses and isinstance(addresses, list):
        lat = addresses[0].get("lat")
        lng = addresses[0].get("lng") or addresses[0].get("lon")
        if lat is not None and lng is not None:
            return safe_float(lat), safe_float(lng)

    # Schema 2.x
    locations = org.get("locations") or []
    if locations and isinstance(locations, list):
        geonames = locations[0].get("geonames_details") or {}
        lat = geonames.get("lat") or geonames.get("latitude")
        lng = geonames.get("lng") or geonames.get("longitude")
        if lat is not None and lng is not None:
            return safe_float(lat), safe_float(lng)
    return None, None


def safe_float(value: Any) -> Optional[float]:
    try:
        if value is None or str(value).lower() == "nan":
            return None
        return float(value)
    except Exception:
        return None


def safe_int(value: Any) -> Optional[int]:
    try:
        if value is None or str(value).lower() == "nan":
            return None
        return int(value)
    except Exception:
        return None


def pct_delta(new: Optional[int], old: Optional[int]) -> Optional[float]:
    if new is None or old in (None, 0):
        return None
    return (new - old) / old


def ror_match(
    session: requests.Session,
    name: str,
    country: Optional[str],
    *,
    ror_client_id: Optional[str] = None,
    timeout: int = 40,
) -> Dict[str, Any]:
    """Devuelve metadatos ROR. Usa affiliation y fallback a query."""
    query_text = f"{name}, {country}" if country else name
    params = {"affiliation": query_text}
    if ror_client_id:
        params["client_id"] = ror_client_id

    data = request_json(session, ROR_BASE_V2, params=params, timeout=timeout)
    chosen_item = None
    if data:
        items = data.get("items") or data.get("results") or []
        if isinstance(items, list) and items:
            chosen_item = next((i for i in items if isinstance(i, dict) and i.get("chosen") is True), None)
            chosen_item = chosen_item or items[0]

    # Fallback a búsqueda general si no hubo match por affiliation.
    if not chosen_item:
        params = {"query": name}
        if ror_client_id:
            params["client_id"] = ror_client_id
        data = request_json(session, ROR_BASE_V1, params=params, timeout=timeout)
        items = (data or {}).get("items") or []
        if isinstance(items, list) and items:
            chosen_item = items[0]

    if not chosen_item:
        return {
            "ror_status_match": "sin_match",
            "ror_id": None,
            "ror_nombre": None,
            "ror_score": None,
            "ror_chosen": None,
            "ror_pais": None,
            "ror_country_code": None,
            "ror_latitud": None,
            "ror_longitud": None,
        }

    org = chosen_item.get("organization") or chosen_item.get("result") or chosen_item
    ror_id = org.get("id") or org.get("ror")
    ror_name = extract_ror_name(org)
    ror_country, ror_country_code = extract_ror_country(org)
    lat, lon = extract_ror_lat_lon(org)

    return {
        "ror_status_match": "ok",
        "ror_id": ror_id,
        "ror_nombre": ror_name,
        "ror_score": chosen_item.get("score") or chosen_item.get("confidence"),
        "ror_chosen": chosen_item.get("chosen"),
        "ror_pais": ror_country,
        "ror_country_code": ror_country_code,
        "ror_latitud": lat,
        "ror_longitud": lon,
    }


def openalex_get_by_ror(
    session: requests.Session,
    ror_id: Optional[str],
    api_key: Optional[str],
    *,
    timeout: int = 40,
) -> Optional[Dict[str, Any]]:
    if not ror_id or not api_key:
        return None
    encoded = urllib.parse.quote(ror_id, safe="")
    url = f"{OPENALEX_BASE}/institutions/{encoded}"
    params = {"api_key": api_key, "select": OPENALEX_SELECT}
    return request_json(session, url, params=params, timeout=timeout)


def openalex_search(
    session: requests.Session,
    name: str,
    country_code: Optional[str],
    api_key: Optional[str],
    *,
    timeout: int = 40,
) -> Optional[Dict[str, Any]]:
    if not api_key:
        return None
    params = {
        "api_key": api_key,
        "search": name,
        "per-page": 5,
        "select": OPENALEX_SELECT,
    }
    if country_code:
        params["filter"] = f"country_code:{country_code}"
    data = request_json(session, f"{OPENALEX_BASE}/institutions", params=params, timeout=timeout)
    results = (data or {}).get("results") or []
    if not results and country_code:
        # Reintento sin filtro por país, útil para territorios o nombres inconsistentes.
        params.pop("filter", None)
        data = request_json(session, f"{OPENALEX_BASE}/institutions", params=params, timeout=timeout)
        results = (data or {}).get("results") or []
    return results[0] if results else None


def counts_for_year(openalex_record: Optional[Dict[str, Any]], year: int) -> Dict[str, Optional[int]]:
    if not openalex_record:
        return {"works": None, "citations": None}
    counts = openalex_record.get("counts_by_year") or []
    for item in counts:
        if safe_int(item.get("year")) == year:
            return {
                "works": safe_int(item.get("works_count")),
                "citations": safe_int(item.get("cited_by_count")),
            }
    return {"works": None, "citations": None}


def flatten_openalex(record: Optional[Dict[str, Any]], target_year: int) -> Dict[str, Any]:
    if not record:
        return {
            "openalex_status_match": "sin_match_o_sin_api_key",
            "openalex_id": None,
            "openalex_nombre": None,
            "openalex_ror": None,
            "openalex_country_code": None,
            "openalex_tipo": None,
            "openalex_homepage": None,
            "openalex_works_total": None,
            "openalex_citas_total": None,
            "openalex_h_index": None,
            "openalex_i10_index": None,
            "openalex_2yr_mean_citedness": None,
            f"openalex_publicaciones_{target_year}": None,
            f"openalex_citas_{target_year}": None,
            f"openalex_publicaciones_{target_year-1}": None,
            f"openalex_citas_{target_year-1}": None,
            "openalex_delta_publicaciones_pct": None,
            "openalex_delta_citas_pct": None,
            "openalex_latitud": None,
            "openalex_longitud": None,
            "openalex_actualizado": None,
            "openalex_url": None,
            "openalex_works_api_url": None,
        }

    stats = record.get("summary_stats") or {}
    geo = record.get("geo") or {}
    current = counts_for_year(record, target_year)
    previous = counts_for_year(record, target_year - 1)

    return {
        "openalex_status_match": "ok",
        "openalex_id": record.get("id"),
        "openalex_nombre": record.get("display_name"),
        "openalex_ror": record.get("ror") or (record.get("ids") or {}).get("ror"),
        "openalex_country_code": record.get("country_code"),
        "openalex_tipo": record.get("type"),
        "openalex_homepage": record.get("homepage_url"),
        "openalex_works_total": safe_int(record.get("works_count")),
        "openalex_citas_total": safe_int(record.get("cited_by_count")),
        "openalex_h_index": safe_int(stats.get("h_index")),
        "openalex_i10_index": safe_int(stats.get("i10_index")),
        "openalex_2yr_mean_citedness": stats.get("2yr_mean_citedness"),
        f"openalex_publicaciones_{target_year}": current["works"],
        f"openalex_citas_{target_year}": current["citations"],
        f"openalex_publicaciones_{target_year-1}": previous["works"],
        f"openalex_citas_{target_year-1}": previous["citations"],
        "openalex_delta_publicaciones_pct": pct_delta(current["works"], previous["works"]),
        "openalex_delta_citas_pct": pct_delta(current["citations"], previous["citations"]),
        "openalex_latitud": geo.get("latitude"),
        "openalex_longitud": geo.get("longitude"),
        "openalex_actualizado": record.get("updated_date"),
        "openalex_url": record.get("id"),
        "openalex_works_api_url": record.get("works_api_url"),
    }


def read_university_dataframe(
    input_path: Path,
    sheet_name: str,
    header_row: Optional[int],
    name_col: Optional[str],
    country_col: Optional[str],
    rank_col: Optional[str],
) -> Tuple[pd.DataFrame, str, Optional[str], Optional[str], int]:
    if header_row is None:
        header_row = find_header_row(input_path, sheet_name)
    df = pd.read_excel(input_path, sheet_name=sheet_name, header=header_row - 1)
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]

    detected_name_col = find_column(df, "name", name_col)
    detected_country_col = None
    detected_rank_col = None
    try:
        detected_country_col = find_column(df, "country", country_col)
    except ValueError:
        logging.warning("No pude detectar columna de país. Continuará sin filtro por país.")
    try:
        detected_rank_col = find_column(df, "rank", rank_col)
    except ValueError:
        logging.warning("No pude detectar columna de ranking. Continuará sin ranking base.")

    df = df[df[detected_name_col].notna()].copy()
    df = df[df[detected_name_col].astype(str).str.strip() != ""].copy()
    return df, detected_name_col, detected_country_col, detected_rank_col, header_row




def normalize_dedupe_text(value: Any) -> str:
    """Normaliza texto para comparar universidades agregadas desde varias hojas."""
    text = normalize_header(value)
    # Limpieza suave de sufijos frecuentes sin intentar hacer matching semántico agresivo.
    for token in [" universidad ", " university "]:
        text = text.replace(token, " ")
    for ch in "'`.,;:()[]{}":
        text = text.replace(ch, " ")
    return " ".join(text.split())


def build_dedupe_key(name: Any, country: Any = None) -> str:
    name_part = normalize_dedupe_text(name)
    country_part = normalize_dedupe_text(country) if country is not None else ""
    return f"{name_part}||{country_part}"


def available_data_sheets(input_path: Path, excluded: Optional[Iterable[str]] = None) -> List[str]:
    """Devuelve hojas que parecen tener una tabla de universidades."""
    excluded_norm = {normalize_header(x) for x in (excluded or [])}
    try:
        xl = pd.ExcelFile(input_path)
    except Exception as exc:
        raise ValueError(f"No pude abrir el Excel: {input_path}. Error: {exc}") from exc

    data_sheets: List[str] = []
    for sheet in xl.sheet_names:
        if normalize_header(sheet) in excluded_norm:
            logging.info("Saltando hoja excluida: %s", sheet)
            continue
        try:
            _ = find_header_row(input_path, sheet)
            data_sheets.append(sheet)
        except Exception:
            logging.info("Saltando hoja sin columna reconocible de universidad: %s", sheet)
    return data_sheets


def standardize_source_dataframe(
    df: pd.DataFrame,
    *,
    sheet_name: str,
    header_row: int,
    name_col: str,
    country_col: Optional[str],
    rank_col: Optional[str],
) -> pd.DataFrame:
    """Agrega columnas estándar para unir hojas con esquemas distintos."""
    out = df.copy()
    out[SOURCE_SHEET_COL] = sheet_name
    out[SOURCE_HEADER_ROW_COL] = header_row
    out[STD_NAME_COL] = out[name_col]
    out[STD_COUNTRY_COL] = out[country_col] if country_col else None
    out[STD_RANK_COL] = out[rank_col] if rank_col else None
    out[DEDUP_KEY_COL] = [
        build_dedupe_key(name, country)
        for name, country in zip(out[STD_NAME_COL], out[STD_COUNTRY_COL])
    ]
    return out


def read_source_dataframes(
    input_path: Path,
    *,
    sheet_name: str,
    all_sheets: bool,
    header_row: Optional[int],
    name_col: Optional[str],
    country_col: Optional[str],
    rank_col: Optional[str],
    dedupe: bool,
) -> Tuple[pd.DataFrame, str, Optional[str], Optional[str], int, List[str]]:
    """
    Lee una hoja o todas las hojas útiles y devuelve un dataframe unificado.

    En modo --all-sheets no se exige que todas las hojas tengan los mismos encabezados:
    cada una se estandariza a tracker_universidad, tracker_pais y tracker_ranking_base.
    """
    excluded = {TRACKER_SHEET, HISTORY_SHEET, "Notas y fuentes", "Notas", "Fuentes"}
    if all_sheets:
        sheets = available_data_sheets(input_path, excluded=excluded)
        if not sheets:
            raise ValueError("No encontró hojas con columnas reconocibles de universidad.")
    else:
        sheets = [sheet_name]

    standardized_frames: List[pd.DataFrame] = []
    detected_header_rows: List[int] = []

    for sheet in sheets:
        try:
            df, detected_name_col, detected_country_col, detected_rank_col, detected_header_row = read_university_dataframe(
                input_path,
                sheet,
                None if all_sheets else header_row,
                name_col,
                country_col,
                rank_col,
            )
        except Exception as exc:
            if all_sheets:
                logging.warning("No pude leer la hoja '%s'; se salta. Motivo: %s", sheet, exc)
                continue
            raise

        logging.info(
            "Hoja '%s': %s filas | encabezado fila %s | universidad=%s | país=%s | ranking=%s",
            sheet,
            len(df),
            detected_header_row,
            detected_name_col,
            detected_country_col or "no detectada",
            detected_rank_col or "no detectado",
        )
        detected_header_rows.append(detected_header_row)
        standardized_frames.append(
            standardize_source_dataframe(
                df,
                sheet_name=sheet,
                header_row=detected_header_row,
                name_col=detected_name_col,
                country_col=detected_country_col,
                rank_col=detected_rank_col,
            )
        )

    if not standardized_frames:
        raise ValueError("No pude construir ninguna tabla de universidades desde el Excel.")

    combined = pd.concat(standardized_frames, ignore_index=True, sort=False)

    if dedupe:
        before = len(combined)
        combined = combined.drop_duplicates(subset=[DEDUP_KEY_COL], keep="first").copy()
        after = len(combined)
        if before != after:
            logging.info("Duplicados exactos eliminados antes de consultar APIs: %s", before - after)

    # A partir de acá usamos columnas estándar, así funciona igual una hoja o muchas.
    common_header_row = detected_header_rows[0] if detected_header_rows else (header_row or 1)
    return combined, STD_NAME_COL, STD_COUNTRY_COL, STD_RANK_COL, common_header_row, sheets


def dedupe_enriched_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Elimina duplicados resueltos por ROR/OpenAlex y luego por nombre+país."""
    if df.empty:
        return df

    out = df.copy()
    before = len(out)

    for col in ["openalex_id", "ror_id"]:
        if col in out.columns:
            has_id = out[col].notna() & (out[col].astype(str).str.strip() != "")
            with_id = out[has_id].drop_duplicates(subset=[col], keep="first")
            without_id = out[~has_id]
            out = pd.concat([with_id, without_id], ignore_index=True, sort=False)

    if DEDUP_KEY_COL in out.columns:
        out = out.drop_duplicates(subset=[DEDUP_KEY_COL], keep="first")

    removed = before - len(out)
    if removed:
        logging.info("Duplicados eliminados después de resolver ROR/OpenAlex: %s", removed)
    return out

def enrich_universities(
    df: pd.DataFrame,
    name_col: str,
    country_col: Optional[str],
    rank_col: Optional[str],
    *,
    target_year: int,
    openalex_api_key: Optional[str],
    ror_client_id: Optional[str],
    cache_path: Path,
    force: bool,
    sleep_seconds: float,
    limit: Optional[int],
    timeout: int,
) -> pd.DataFrame:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "UniversidadRankingTracker/1.0 (contact: local-script)",
        "Accept": "application/json",
    })

    cache = load_cache(cache_path)
    items_cache = cache.setdefault("items", {})

    rows: List[Dict[str, Any]] = []
    total = len(df) if limit is None else min(len(df), limit)

    if not openalex_api_key:
        logging.warning(
            "No se encontró OPENALEX_API_KEY ni --openalex-api-key. "
            "Se consultará ROR, pero las métricas OpenAlex quedarán vacías."
        )

    for idx, (_, row) in enumerate(df.iterrows(), start=1):
        if limit is not None and idx > limit:
            break

        name = str(row.get(name_col, "")).strip()
        country = str(row.get(country_col, "")).strip() if country_col and pd.notna(row.get(country_col)) else None
        country_code = country_to_alpha2(country) if country else None
        rank = row.get(rank_col) if rank_col else None

        logging.info("[%s/%s] %s (%s)", idx, total, name, country or "sin país")

        cache_key = f"{name}||{country or ''}||{target_year}"
        if not force and cache_key in items_cache:
            enriched = items_cache[cache_key]
        else:
            ror_data = ror_match(
                session,
                name,
                country,
                ror_client_id=ror_client_id,
                timeout=timeout,
            )
            openalex_record = openalex_get_by_ror(
                session,
                ror_data.get("ror_id"),
                openalex_api_key,
                timeout=timeout,
            )
            if openalex_record is None:
                openalex_record = openalex_search(
                    session,
                    name,
                    country_code or ror_data.get("ror_country_code"),
                    openalex_api_key,
                    timeout=timeout,
                )

            openalex_data = flatten_openalex(openalex_record, target_year)
            enriched = {
                **ror_data,
                **openalex_data,
                "tracker_fecha_utc": now_iso(),
                "tracker_target_year": target_year,
            }
            items_cache[cache_key] = enriched
            save_cache(cache_path, cache)
            if sleep_seconds > 0:
                time.sleep(sleep_seconds)

        base = row.to_dict()
        base["ranking_base_del_excel"] = rank
        base["pais_alpha2_detectado"] = country_code
        rows.append({**base, **enriched})

    return pd.DataFrame(rows)


def make_history(df: pd.DataFrame, name_col: str, country_col: Optional[str], rank_col: Optional[str], target_year: int) -> pd.DataFrame:
    timestamp = now_iso()
    history_columns = [
        "timestamp_utc",
        "universidad",
        "pais",
        "ranking_base",
        "target_year",
        "ror_id",
        "openalex_id",
        "openalex_works_total",
        "openalex_citas_total",
        "openalex_h_index",
        "openalex_i10_index",
        f"openalex_publicaciones_{target_year}",
        f"openalex_citas_{target_year}",
    ]
    if df.empty:
        return pd.DataFrame(columns=history_columns)

    cols = {
        "timestamp_utc": [timestamp] * len(df),
        "universidad": df.get(name_col),
        "pais": df.get(country_col) if country_col else None,
        "ranking_base": df.get(rank_col) if rank_col else df.get("ranking_base_del_excel"),
        "target_year": [target_year] * len(df),
        "ror_id": df.get("ror_id"),
        "openalex_id": df.get("openalex_id"),
        "openalex_works_total": df.get("openalex_works_total"),
        "openalex_citas_total": df.get("openalex_citas_total"),
        "openalex_h_index": df.get("openalex_h_index"),
        "openalex_i10_index": df.get("openalex_i10_index"),
        f"openalex_publicaciones_{target_year}": df.get(f"openalex_publicaciones_{target_year}"),
        f"openalex_citas_{target_year}": df.get(f"openalex_citas_{target_year}"),
    }
    return pd.DataFrame(cols, columns=history_columns)


def df_column(df: pd.DataFrame, column: Optional[str], default: Any = None) -> pd.Series:
    if column and column in df.columns:
        return df[column]
    return pd.Series([default] * len(df), index=df.index)


def make_institutions_master(df: pd.DataFrame, name_col: str, country_col: Optional[str]) -> pd.DataFrame:
    columns = [
        "institution_key",
        "universidad_original",
        "pais_original",
        "pais_alpha2_detectado",
        "ror_id",
        "ror_nombre",
        "ror_pais",
        "ror_country_code",
        "openalex_id",
        "openalex_nombre",
        "openalex_country_code",
        "openalex_tipo",
        "homepage",
        "latitud",
        "longitud",
        "match_status",
        "match_confidence_ror",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)

    out = pd.DataFrame({
        "institution_key": [build_dedupe_key(name, country) for name, country in zip(df_column(df, name_col), df_column(df, country_col))],
        "universidad_original": df_column(df, name_col),
        "pais_original": df_column(df, country_col),
        "pais_alpha2_detectado": df_column(df, "pais_alpha2_detectado"),
        "ror_id": df_column(df, "ror_id"),
        "ror_nombre": df_column(df, "ror_nombre"),
        "ror_pais": df_column(df, "ror_pais"),
        "ror_country_code": df_column(df, "ror_country_code"),
        "openalex_id": df_column(df, "openalex_id"),
        "openalex_nombre": df_column(df, "openalex_nombre"),
        "openalex_country_code": df_column(df, "openalex_country_code"),
        "openalex_tipo": df_column(df, "openalex_tipo"),
        "homepage": df_column(df, "openalex_homepage"),
        "latitud": df_column(df, "openalex_latitud").combine_first(df_column(df, "ror_latitud")),
        "longitud": df_column(df, "openalex_longitud").combine_first(df_column(df, "ror_longitud")),
        "match_status": [
            "ok" if ror == "ok" and openalex == "ok" else "revisar"
            for ror, openalex in zip(df_column(df, "ror_status_match"), df_column(df, "openalex_status_match"))
        ],
        "match_confidence_ror": df_column(df, "ror_score"),
    })
    return out.drop_duplicates(subset=["institution_key", "ror_id", "openalex_id"], keep="first")


def make_rankings_history_master(df: pd.DataFrame, name_col: str, country_col: Optional[str], rank_col: Optional[str]) -> pd.DataFrame:
    columns = [
        "source",
        "edition",
        "universidad",
        "pais",
        "ranking_base",
        "hoja_origen",
        "ror_id",
        "openalex_id",
        "observacion",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)
    ranking_series = df_column(df, rank_col).combine_first(df_column(df, "ranking_base_del_excel"))
    out = pd.DataFrame({
        "source": ["excel_base"] * len(df),
        "edition": [None] * len(df),
        "universidad": df_column(df, name_col),
        "pais": df_column(df, country_col),
        "ranking_base": ranking_series,
        "hoja_origen": df_column(df, SOURCE_SHEET_COL),
        "ror_id": df_column(df, "ror_id"),
        "openalex_id": df_column(df, "openalex_id"),
        "observacion": ["Ranking conservado desde el Excel de entrada."] * len(df),
    })
    return out


def make_openalex_metrics_master(df: pd.DataFrame, name_col: str, country_col: Optional[str], target_year: int) -> pd.DataFrame:
    columns = [
        "universidad",
        "pais",
        "ror_id",
        "openalex_id",
        "openalex_nombre",
        "works_total",
        "citas_total",
        "h_index",
        "i10_index",
        "mean_citedness_2yr",
        f"publicaciones_{target_year}",
        f"citas_{target_year}",
        f"publicaciones_{target_year - 1}",
        f"citas_{target_year - 1}",
        "delta_publicaciones_pct",
        "delta_citas_pct",
        "openalex_actualizado",
        "openalex_url",
        "works_api_url",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame({
        "universidad": df_column(df, name_col),
        "pais": df_column(df, country_col),
        "ror_id": df_column(df, "ror_id"),
        "openalex_id": df_column(df, "openalex_id"),
        "openalex_nombre": df_column(df, "openalex_nombre"),
        "works_total": df_column(df, "openalex_works_total"),
        "citas_total": df_column(df, "openalex_citas_total"),
        "h_index": df_column(df, "openalex_h_index"),
        "i10_index": df_column(df, "openalex_i10_index"),
        "mean_citedness_2yr": df_column(df, "openalex_2yr_mean_citedness"),
        f"publicaciones_{target_year}": df_column(df, f"openalex_publicaciones_{target_year}"),
        f"citas_{target_year}": df_column(df, f"openalex_citas_{target_year}"),
        f"publicaciones_{target_year - 1}": df_column(df, f"openalex_publicaciones_{target_year - 1}"),
        f"citas_{target_year - 1}": df_column(df, f"openalex_citas_{target_year - 1}"),
        "delta_publicaciones_pct": df_column(df, "openalex_delta_publicaciones_pct"),
        "delta_citas_pct": df_column(df, "openalex_delta_citas_pct"),
        "openalex_actualizado": df_column(df, "openalex_actualizado"),
        "openalex_url": df_column(df, "openalex_url"),
        "works_api_url": df_column(df, "openalex_works_api_url"),
    })


def make_pending_matches(df: pd.DataFrame, name_col: str, country_col: Optional[str]) -> pd.DataFrame:
    columns = [
        "universidad",
        "pais",
        "ror_status_match",
        "ror_score",
        "ror_id",
        "ror_nombre",
        "openalex_status_match",
        "openalex_id",
        "openalex_nombre",
        "motivo_revision",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)

    rows: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        reasons: List[str] = []
        if row.get("ror_status_match") != "ok":
            reasons.append("sin_match_ror")
        if row.get("openalex_status_match") != "ok":
            reasons.append("sin_match_openalex")
        score = row.get("ror_score")
        if pd.notna(score):
            try:
                if float(score) < 0.8:
                    reasons.append("score_ror_bajo")
            except (TypeError, ValueError):
                pass
        if not reasons:
            continue
        rows.append({
            "universidad": row.get(name_col),
            "pais": row.get(country_col) if country_col else None,
            "ror_status_match": row.get("ror_status_match"),
            "ror_score": row.get("ror_score"),
            "ror_id": row.get("ror_id"),
            "ror_nombre": row.get("ror_nombre"),
            "openalex_status_match": row.get("openalex_status_match"),
            "openalex_id": row.get("openalex_id"),
            "openalex_nombre": row.get("openalex_nombre"),
            "motivo_revision": ", ".join(reasons),
        })
    return pd.DataFrame(rows, columns=columns)


def make_quality_control(df: pd.DataFrame, name_col: str, country_col: Optional[str], rank_col: Optional[str]) -> pd.DataFrame:
    checks: List[Dict[str, Any]] = []

    def add_check(check: str, severity: str, count: int, note: str) -> None:
        checks.append({
            "check": check,
            "severity": severity,
            "count": int(count),
            "note": note,
            "timestamp_utc": now_iso(),
        })

    if df.empty:
        add_check("dataset_vacio", "error", 1, "No hay universidades procesadas.")
        return pd.DataFrame(checks)

    add_check("filas_procesadas", "info", len(df), "Total de universidades en el tracker enriquecido.")
    add_check("sin_ror_id", "warning", int(df_column(df, "ror_id").isna().sum()), "Instituciones sin identificador ROR.")
    add_check("sin_openalex_id", "warning", int(df_column(df, "openalex_id").isna().sum()), "Instituciones sin identificador OpenAlex.")
    ranking_series = df_column(df, rank_col).combine_first(df_column(df, "ranking_base_del_excel"))
    add_check("sin_ranking_base", "info", int(ranking_series.isna().sum()), "Instituciones sin ranking base en el Excel de entrada.")
    if country_col:
        add_check("sin_pais", "warning", int(df_column(df, country_col).isna().sum()), "Instituciones sin país detectado.")
    duplicate_keys = [
        build_dedupe_key(name, country)
        for name, country in zip(df_column(df, name_col), df_column(df, country_col))
    ]
    add_check("duplicados_nombre_pais", "warning", int(pd.Series(duplicate_keys).duplicated().sum()), "Duplicados por nombre normalizado + país.")
    if "ror_id" in df.columns:
        add_check("duplicados_ror_id", "warning", int(df["ror_id"].dropna().duplicated().sum()), "Duplicados con el mismo ROR ID.")
    if "openalex_id" in df.columns:
        add_check("duplicados_openalex_id", "warning", int(df["openalex_id"].dropna().duplicated().sum()), "Duplicados con el mismo OpenAlex ID.")
    pending = make_pending_matches(df, name_col, country_col)
    add_check("matches_pendientes", "warning", len(pending), "Instituciones que requieren revisión manual de match.")
    return pd.DataFrame(checks)


def read_existing_history(output_path: Path) -> pd.DataFrame:
    if not output_path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(output_path, sheet_name=HISTORY_SHEET)
    except Exception:
        return pd.DataFrame()


def sanitize_sheet_title(title: str) -> str:
    invalid = set("[]:*?/\\")
    cleaned = "".join("_" if ch in invalid else ch for ch in title)
    return cleaned[:31]


def write_dataframe_sheet(wb, sheet_name: str, df: pd.DataFrame, title: str, note: str) -> None:
    sheet_name = sanitize_sheet_title(sheet_name)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    max_col = max(1, len(df.columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="1F4E78")
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(2, 1, note)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)

    start_row = 4
    for r_offset, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        r_idx = start_row + r_offset
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            if r_idx == start_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="5B9BD5")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    end_row = start_row + len(df)
    end_col_letter = get_column_letter(max_col)
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{end_col_letter}{end_row}"

    if len(df) > 0 and max_col > 0:
        table_ref = f"A{start_row}:{end_col_letter}{end_row}"
        table_name = "tbl_" + "".join(ch if ch.isalnum() else "_" for ch in sheet_name)[:20]
        try:
            tab = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)
        except Exception:
            pass

    # Ajuste simple de anchos, con límites razonables.
    for col_idx, col_name in enumerate(df.columns, start=1):
        series = df[col_name].astype(str).head(200) if len(df) else []
        max_len = max([len(str(col_name))] + [len(str(x)) for x in series]) if len(df) else len(str(col_name))
        width = min(max(max_len + 2, 10), 42)
        if "url" in str(col_name).lower() or "maps" in str(col_name).lower() or "api" in str(col_name).lower():
            width = 32
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_output_workbook(
    input_path: Path,
    output_path: Path,
    tracker_df: pd.DataFrame,
    history_df: pd.DataFrame,
    *,
    name_col: str,
    country_col: Optional[str],
    rank_col: Optional[str],
    target_year: int,
) -> None:
    old_history = read_existing_history(output_path)
    new_history = make_history(tracker_df, name_col, country_col, rank_col, target_year)
    if not old_history.empty:
        history_df = pd.concat([old_history, new_history], ignore_index=True)
    else:
        history_df = new_history

    institutions_df = make_institutions_master(tracker_df, name_col, country_col)
    rankings_history_df = make_rankings_history_master(tracker_df, name_col, country_col, rank_col)
    openalex_metrics_df = make_openalex_metrics_master(tracker_df, name_col, country_col, target_year)
    pending_matches_df = make_pending_matches(tracker_df, name_col, country_col)
    quality_df = make_quality_control(tracker_df, name_col, country_col, rank_col)

    wb = load_workbook(input_path)
    write_dataframe_sheet(
        wb,
        INSTITUTIONS_SHEET,
        institutions_df,
        "Instituciones normalizadas",
        "Catálogo maestro de instituciones con identificadores ROR/OpenAlex y estado de matching.",
    )
    write_dataframe_sheet(
        wb,
        RANKINGS_HISTORY_SHEET,
        rankings_history_df,
        "Rankings históricos",
        "Ranking base conservado desde el Excel de entrada; puede combinarse con rankings oficiales públicos recolectados por --scrape-rankings.",
    )
    write_dataframe_sheet(
        wb,
        OPENALEX_METRICS_SHEET,
        openalex_metrics_df,
        "Métricas OpenAlex",
        "Métricas académicas separadas en formato analítico para comparación y reportes.",
    )
    write_dataframe_sheet(
        wb,
        PENDING_MATCHES_SHEET,
        pending_matches_df,
        "Matches pendientes",
        "Instituciones que requieren revisión manual por falta de ROR/OpenAlex o baja confianza del match.",
    )
    write_dataframe_sheet(
        wb,
        QUALITY_SHEET,
        quality_df,
        "Control de calidad",
        "Resumen de validaciones de datos, duplicados, faltantes y matches pendientes.",
    )
    write_dataframe_sheet(
        wb,
        TRACKER_SHEET,
        tracker_df,
        "Tracker de universidades - ROR + OpenAlex",
        "Ranking QS/THE/ARWU no se actualiza en tiempo real: esta hoja conserva el ranking base y suma métricas académicas actualizables desde ROR/OpenAlex.",
    )
    write_dataframe_sheet(
        wb,
        HISTORY_SHEET,
        history_df,
        "Historial de métricas académicas",
        "Cada ejecución agrega una foto temporal de las métricas para comparar evolución.",
    )
    wb.save(output_path)


RANKING_SOURCE_CONFIG = {
    "qs": {
        "name": "QS World University Rankings",
        "latest_year": 2026,
        "url_template": "https://www.topuniversities.com/world-university-rankings/{year}",
        "fallback_url": "https://www.topuniversities.com/world-university-rankings",
    },
    "the": {
        "name": "Times Higher Education World University Rankings",
        "latest_year": 2026,
        "url_template": "https://www.timeshighereducation.com/world-university-rankings/{year}/world-ranking",
        "fallback_url": "https://www.timeshighereducation.com/world-university-rankings/latest/world-ranking",
    },
    "arwu": {
        "name": "ShanghaiRanking ARWU",
        "latest_year": 2025,
        "url_template": "https://www.shanghairanking.com/rankings/arwu/{year}",
        "fallback_url": "https://www.shanghairanking.com/rankings/arwu/2025",
    },
}


def clean_html_text(value: str) -> str:
    text = re.sub(r"<script\b.*?</script>", " ", value, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"<style\b.*?</style>", " ", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"<[^>]+>", " ", text)
    text = text.replace("&nbsp;", " ").replace("&amp;", "&")
    text = text.replace("&#39;", "'").replace("&quot;", '"')
    return " ".join(text.split())


def find_public_data_links(base_url: str, html: str) -> List[str]:
    links: List[str] = []
    for match in re.finditer(r"""(?:href|src)=["']([^"']+\.(?:json|csv|xlsx|xls)(?:\?[^"']*)?)["']""", html, re.IGNORECASE):
        url = urllib.parse.urljoin(base_url, match.group(1))
        if url not in links:
            links.append(url)
    return links


def flatten_json_records(value: Any) -> List[Dict[str, Any]]:
    candidates: List[List[Dict[str, Any]]] = []

    def visit(node: Any) -> None:
        if isinstance(node, list):
            dict_items = [x for x in node if isinstance(x, dict)]
            if len(dict_items) >= 3:
                keys = {normalize_header(k) for item in dict_items[:10] for k in item.keys()}
                if keys & {"rank", "rank display", "rank_order", "name", "institution", "university", "scores overall"}:
                    candidates.append(dict_items)
            for item in node:
                visit(item)
        elif isinstance(node, dict):
            for item in node.values():
                visit(item)

    visit(value)
    if not candidates:
        return []
    return max(candidates, key=len)


def dataframe_from_public_link(session: requests.Session, url: str, timeout: int) -> pd.DataFrame:
    resp = session.get(url, timeout=timeout)
    resp.raise_for_status()
    lower_url = url.lower()
    content_type = resp.headers.get("content-type", "").lower()
    if ".json" in lower_url or "json" in content_type:
        records = flatten_json_records(resp.json())
        return pd.DataFrame(records)
    if ".csv" in lower_url or "csv" in content_type:
        return pd.read_csv(io.BytesIO(resp.content))
    if any(ext in lower_url for ext in (".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(resp.content))
    return pd.DataFrame()


def normalize_scraped_dataframe(df: pd.DataFrame, source: str, ranking_name: str, year: int, source_url: str) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out.insert(0, "source", source)
    out.insert(1, "ranking_name", ranking_name)
    out.insert(2, "edition_year", year)
    out.insert(3, "source_url", source_url)
    out.insert(4, "retrieved_at", now_iso())
    return out


def scrape_arwu_public_html(session: requests.Session, year: int, timeout: int) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    config = RANKING_SOURCE_CONFIG["arwu"]
    url = config["url_template"].format(year=year)
    resp = session.get(url, timeout=timeout)
    status = {
        "source": "arwu",
        "ranking_name": config["name"],
        "edition_year": year,
        "source_url": url,
        "status_code": resp.status_code,
        "status": "ok",
        "rows": 0,
        "note": "Tabla pública HTML renderizada por ShanghaiRanking.",
        "retrieved_at": now_iso(),
    }
    resp.raise_for_status()
    rows: List[Dict[str, Any]] = []
    for row_html in re.findall(r"<tr\b[^>]*data-v-c5e69b9e[^>]*>(.*?)</tr>", resp.text, flags=re.IGNORECASE | re.DOTALL):
        rank_match = re.search(r'<div class="ranking"[^>]*>(.*?)</div>', row_html, flags=re.IGNORECASE | re.DOTALL)
        name_match = re.search(r'<span class="univ-name"[^>]*>(.*?)</span>', row_html, flags=re.IGNORECASE | re.DOTALL)
        country_match = re.search(r'<div class="location"[^>]*>.*?</img>\s*(.*?)</div>', row_html, flags=re.IGNORECASE | re.DOTALL)
        link_match = re.search(r'<a href="([^"]+)"[^>]*>\s*<span class="univ-name"', row_html, flags=re.IGNORECASE | re.DOTALL)
        cells = [clean_html_text(cell) for cell in re.findall(r"<td\b[^>]*>(.*?)</td>", row_html, flags=re.IGNORECASE | re.DOTALL)]
        if not name_match:
            continue
        metric_cells = cells[3:] if len(cells) >= 4 else []
        record: Dict[str, Any] = {
            "rank": clean_html_text(rank_match.group(1)) if rank_match else (cells[0] if cells else None),
            "university": clean_html_text(name_match.group(1)),
            "country": clean_html_text(country_match.group(1)) if country_match else None,
            "national_rank": cells[2] if len(cells) >= 3 else None,
            "institution_url": urllib.parse.urljoin(url, link_match.group(1)) if link_match else None,
        }
        for key, value in zip(["total_score", "alumni", "award", "hici", "n_s", "pub", "pcp"], metric_cells):
            record[key] = value
        rows.append(record)
    status["rows"] = len(rows)
    if rows and len(rows) < 1000:
        status["status"] = "partial_public_html"
        status["note"] = (
            "La página oficial informa paginación, pero el HTML público recuperado contiene "
            f"{len(rows)} filas renderizadas. No se intenta eludir controles ni usar datos no publicados."
        )
    return normalize_scraped_dataframe(pd.DataFrame(rows), "arwu", config["name"], year, url), status


def scrape_official_data_url(session: requests.Session, source: str, year: int, data_url: str, timeout: int) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    config = RANKING_SOURCE_CONFIG[source]
    status = {
        "source": source,
        "ranking_name": config["name"],
        "edition_year": year,
        "source_url": data_url,
        "status_code": None,
        "status": "not_started",
        "rows": 0,
        "note": "URL oficial de datos provista explícitamente por el usuario.",
        "retrieved_at": now_iso(),
    }
    try:
        df = dataframe_from_public_link(session, data_url, timeout)
    except Exception as exc:
        status["status"] = "data_url_failed"
        status["note"] = f"No se pudo leer la URL oficial provista: {exc}"
        return pd.DataFrame(), status
    if df.empty:
        status["status"] = "data_url_empty"
        status["note"] = "La URL oficial provista respondió, pero no produjo filas tabulares."
        return pd.DataFrame(), status
    out = normalize_scraped_dataframe(df, source, config["name"], year, data_url)
    status["status"] = "ok"
    status["rows"] = len(out)
    return out, status


def scrape_generic_official_source(session: requests.Session, source: str, year: int, timeout: int) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    config = RANKING_SOURCE_CONFIG[source]
    url = config["url_template"].format(year=year)
    status = {
        "source": source,
        "ranking_name": config["name"],
        "edition_year": year,
        "source_url": url,
        "status_code": None,
        "status": "not_started",
        "rows": 0,
        "note": "",
        "retrieved_at": now_iso(),
    }
    headers = {
        "User-Agent": "universidad-ranking-tracker/1.2 (+public official data collection; contact: repository user)",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    try:
        resp = session.get(url, headers=headers, timeout=timeout)
        status["status_code"] = resp.status_code
        if resp.status_code == 404 and config.get("fallback_url"):
            url = str(config["fallback_url"])
            status["source_url"] = url
            resp = session.get(url, headers=headers, timeout=timeout)
            status["status_code"] = resp.status_code
        resp.raise_for_status()
    except requests.RequestException as exc:
        status["status"] = "blocked_or_unavailable"
        status["note"] = f"No se pudo acceder a la página oficial pública: {exc}"
        return pd.DataFrame(), status

    html = resp.text
    if "Just a moment" in html and "Cloudflare" in html:
        status["status"] = "blocked_by_provider"
        status["note"] = "La fuente oficial respondió con desafío Cloudflare; no se intenta eludirlo."
        return pd.DataFrame(), status

    frames: List[pd.DataFrame] = []
    data_links = find_public_data_links(url, html)
    for link in data_links:
        try:
            link_df = dataframe_from_public_link(session, link, timeout)
            if not link_df.empty:
                frames.append(normalize_scraped_dataframe(link_df, source, config["name"], year, link))
        except Exception as exc:
            logging.debug("No pude leer enlace público %s: %s", link, exc)

    try:
        html_tables = pd.read_html(io.StringIO(html))
        for table_df in html_tables:
            if len(table_df) >= 3:
                frames.append(normalize_scraped_dataframe(table_df, source, config["name"], year, url))
    except Exception as exc:
        logging.debug("No pude leer tablas HTML en %s: %s", url, exc)

    if frames:
        df = pd.concat(frames, ignore_index=True, sort=False)
        status["status"] = "ok"
        status["rows"] = len(df)
        status["note"] = f"Datos obtenidos desde enlaces/tablas públicos oficiales detectados: {len(data_links)} enlaces."
        return df, status

    status["status"] = "no_public_table_detected"
    status["note"] = (
        "La página oficial respondió, pero no expuso una tabla HTML ni enlaces JSON/CSV/XLSX públicos detectables. "
        "Si la fuente ofrece descarga oficial manual, pasá ese archivo como Excel base o agregá el enlace oficial."
    )
    return pd.DataFrame(), status


def parse_ranking_data_urls(value: str) -> Dict[str, str]:
    urls: Dict[str, str] = {}
    for item in str(value or "").split(","):
        if not item.strip():
            continue
        if "=" not in item:
            raise ValueError("Cada --ranking-data-urls debe tener formato fuente=url o fuente:año=url.")
        source, url = item.split("=", 1)
        urls[source.strip().lower()] = url.strip()
    return urls


def ranking_data_url_for_year(data_urls: Dict[str, str], source: str, year: int) -> Optional[str]:
    return data_urls.get(f"{source}:{year}") or data_urls.get(source)


def load_ranking_config(config_path: Optional[str]) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, str]]:
    config = {key: dict(value) for key, value in RANKING_SOURCE_CONFIG.items()}
    data_urls: Dict[str, str] = {}
    if not config_path:
        return config, data_urls

    path = Path(config_path).expanduser().resolve()
    raw = json.loads(path.read_text(encoding="utf-8"))
    sources = raw.get("sources", raw)
    if not isinstance(sources, dict):
        raise ValueError("El archivo --ranking-config debe contener un objeto JSON o una clave 'sources'.")

    for source, source_config in sources.items():
        source_key = str(source).strip().lower()
        if not isinstance(source_config, dict):
            raise ValueError(f"La configuración de '{source_key}' debe ser un objeto JSON.")
        merged = dict(config.get(source_key, {}))
        for key in ("name", "latest_year", "url_template", "fallback_url"):
            if key in source_config:
                merged[key] = source_config[key]
        config[source_key] = merged

        source_data_urls = source_config.get("data_urls") or {}
        if isinstance(source_data_urls, str):
            data_urls[source_key] = source_data_urls
        elif isinstance(source_data_urls, dict):
            for year, url in source_data_urls.items():
                data_urls[f"{source_key}:{year}"] = str(url)
        elif source_data_urls:
            raise ValueError(f"'data_urls' de '{source_key}' debe ser string u objeto año->url.")
    return config, data_urls


def scrape_public_rankings(
    sources: List[str],
    years: List[int],
    timeout: int,
    data_urls: Optional[Dict[str, str]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    session = requests.Session()
    data_urls = data_urls or {}
    frames: List[pd.DataFrame] = []
    statuses: List[Dict[str, Any]] = []
    for source in sources:
        if source not in RANKING_SOURCE_CONFIG:
            statuses.append({
                "source": source,
                "ranking_name": None,
                "edition_year": None,
                "source_url": None,
                "status_code": None,
                "status": "unknown_source",
                "rows": 0,
                "note": f"Fuente no soportada. Usá una de: {', '.join(RANKING_SOURCE_CONFIG)}",
                "retrieved_at": now_iso(),
            })
            continue
        selected_years = years or [int(RANKING_SOURCE_CONFIG[source]["latest_year"])]
        for year in selected_years:
            logging.info("Scraping oficial público: %s %s", source.upper(), year)
            data_url = ranking_data_url_for_year(data_urls, source, year)
            if data_url:
                df, status = scrape_official_data_url(session, source, year, data_url, timeout)
            elif source == "arwu":
                df, status = scrape_arwu_public_html(session, year, timeout)
            else:
                df, status = scrape_generic_official_source(session, source, year, timeout)
            if not df.empty:
                frames.append(df)
            statuses.append(status)
    ranking_df = pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()
    status_df = pd.DataFrame(statuses)
    return ranking_df, status_df


def write_rankings_workbook(output_path: Path, ranking_df: pd.DataFrame, status_df: pd.DataFrame) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        ranking_df.to_excel(writer, sheet_name=RANKINGS_SHEET, index=False)
        status_df.to_excel(writer, sheet_name=RANKINGS_SOURCES_SHEET, index=False)


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Actualiza métricas OpenAlex/ROR y recolecta rankings oficiales públicos cuando están disponibles.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input", help="Ruta del Excel de entrada.")
    parser.add_argument("--output", default="universidades_tracker_actualizado.xlsx", help="Ruta del Excel de salida.")
    parser.add_argument("--scrape-rankings", action="store_true", help="Recolecta rankings oficiales públicos y genera un Excel independiente.")
    parser.add_argument("--ranking-sources", default="qs,the,arwu", help="Fuentes a consultar separadas por coma: qs,the,arwu.")
    parser.add_argument("--ranking-years", default="", help="Años a consultar separados por coma. Si se omite, usa el último año configurado por fuente.")
    parser.add_argument("--ranking-config", default="", help="JSON opcional para actualizar fuentes, últimos años y URLs oficiales directas sin tocar el código.")
    parser.add_argument("--ranking-data-urls", default="", help="URLs oficiales de datos separadas por coma con formato fuente=url o fuente:año=url. Ejemplo: qs:2027=https://...xlsx")
    parser.add_argument("--rankings-output", default="rankings_oficiales_publicos.xlsx", help="Excel de salida para --scrape-rankings.")
    parser.add_argument("--sheet", default="Top 200 QS 2026", help="Hoja a leer del Excel cuando no se usa --all-sheets.")
    parser.add_argument("--all-sheets", action="store_true", help="Procesa automáticamente todas las hojas que tengan columnas reconocibles de universidad.")
    parser.add_argument("--no-dedupe", action="store_true", help="No elimina duplicados entre hojas. Por defecto se deduplica por nombre+país y luego por ROR/OpenAlex.")
    parser.add_argument("--header-row", type=int, default=None, help="Fila de encabezados. Si se omite, se intenta detectar automáticamente.")
    parser.add_argument("--name-col", default=None, help="Nombre exacto de la columna universidad.")
    parser.add_argument("--country-col", default=None, help="Nombre exacto de la columna país.")
    parser.add_argument("--rank-col", default=None, help="Nombre exacto de la columna ranking.")
    parser.add_argument("--target-year", type=int, default=dt.date.today().year - 1, help="Año para comparar métricas por año; recomendado: último año completo.")
    parser.add_argument("--openalex-api-key", default=os.getenv("OPENALEX_API_KEY"), help="API key de OpenAlex. También se puede usar variable OPENALEX_API_KEY.")
    parser.add_argument("--ror-client-id", default=os.getenv("ROR_CLIENT_ID"), help="Client ID de ROR si tenés uno. También variable ROR_CLIENT_ID.")
    parser.add_argument("--cache", default="universidades_tracker_cache.json", help="Archivo JSON de caché para no repetir consultas.")
    parser.add_argument("--force", action="store_true", help="Ignora caché y vuelve a consultar APIs.")
    parser.add_argument("--sleep", type=float, default=0.15, help="Pausa entre universidades para ser amable con las APIs.")
    parser.add_argument("--limit", type=int, default=None, help="Procesa solo N universidades, útil para pruebas.")
    parser.add_argument("--timeout", type=int, default=40, help="Timeout HTTP por consulta, en segundos.")
    parser.add_argument("--verbose", action="store_true", help="Muestra logs detallados.")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    global RANKING_SOURCE_CONFIG

    args = parse_args(argv)
    setup_logging(args.verbose)

    if args.scrape_rankings:
        config, config_data_urls = load_ranking_config(args.ranking_config)
        RANKING_SOURCE_CONFIG = config
        sources = [s.strip().lower() for s in str(args.ranking_sources).split(",") if s.strip()]
        years = [int(y.strip()) for y in str(args.ranking_years).split(",") if y.strip()]
        data_urls = {**config_data_urls, **parse_ranking_data_urls(args.ranking_data_urls)}
        output_path = Path(args.rankings_output).expanduser().resolve()
        ranking_df, status_df = scrape_public_rankings(sources, years, args.timeout, data_urls)
        logging.info("Escribiendo rankings oficiales públicos: %s", output_path)
        write_rankings_workbook(output_path, ranking_df, status_df)
        logging.info("Listo: %s | filas ranking=%s | fuentes=%s", output_path, len(ranking_df), len(status_df))
        return 0

    if not args.input:
        logging.error("Falta --input. Para scraping de rankings usá --scrape-rankings.")
        return 2

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    cache_path = Path(args.cache).expanduser().resolve()

    if not input_path.exists():
        logging.error("No existe el archivo de entrada: %s", input_path)
        return 2

    logging.info("Leyendo Excel: %s", input_path)
    df, name_col, country_col, rank_col, header_row, processed_sheets = read_source_dataframes(
        input_path,
        sheet_name=args.sheet,
        all_sheets=args.all_sheets,
        header_row=args.header_row,
        name_col=args.name_col,
        country_col=args.country_col,
        rank_col=args.rank_col,
        dedupe=not args.no_dedupe,
    )
    logging.info("Hojas procesadas: %s", ", ".join(processed_sheets))
    logging.info("Total de universidades a procesar: %s", len(df))
    logging.info("Columnas estándar usadas: universidad=%s | país=%s | ranking=%s", name_col, country_col, rank_col)

    tracker_df = enrich_universities(
        df,
        name_col,
        country_col,
        rank_col,
        target_year=args.target_year,
        openalex_api_key=args.openalex_api_key,
        ror_client_id=args.ror_client_id,
        cache_path=cache_path,
        force=args.force,
        sleep_seconds=args.sleep,
        limit=args.limit,
        timeout=args.timeout,
    )

    if not args.no_dedupe:
        tracker_df = dedupe_enriched_dataframe(tracker_df)

    logging.info("Escribiendo Excel actualizado: %s", output_path)
    write_output_workbook(
        input_path,
        output_path,
        tracker_df,
        pd.DataFrame(),
        name_col=name_col,
        country_col=country_col,
        rank_col=rank_col,
        target_year=args.target_year,
    )
    logging.info("Listo: %s", output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

